from openpyxl import load_workbook


# Load the existing workbook
workbook = load_workbook("profiles.xlsx")

# Select the sheet from the workbook
sheet = workbook["Sheet"]
number_column = 1  # Column number for candidate numbers
link_column = 2   # Column number for candidate links

# Initialize an empty dictionary to store candidate names and links
existing_candidates = set()

# Populate the set with existing candidate names and links
for row in sheet.iter_rows(min_row=2, min_col=link_column, max_col=link_column, values_only=True):
    existing_candidates.add(row[0])

# Find the next available candidate number
number_values = [row[0].value for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=number_column, max_col=number_column) if row[0].value is not None]
next_candidate_number = max(number_values, default=0) + 1

# Find the last used row in the sheet
last_used_row = sheet.max_row
#last_used_row = 16 
# Set the starting row for the next entry
starting_row = last_used_row + 1

# Loop to collect candidate names and links
while True:
    # Ask the user for candidate name (exit loop if "exit" is entered)
    user_name = input("Please enter candidate name (enter 'exit' to quit): ")
    if user_name.lower() == "exit":
        break

    if user_name in existing_candidates:
        print("Candidate name already exists. Please enter a unique name.")
        continue
    
    # Ask the user for candidate link
    user_link = input("Please enter a candidate link: ")
    
    if user_link in [row[0].value for row in sheet.iter_rows(min_row=2, max_row=last_used_row, min_col=link_column, max_col=link_column)]:
        print("Candidate link already exists. Please enter a unique link.")
        continue
    
    # Write the candidate number to the sheet
    cell_number = sheet.cell(row=starting_row, column=number_column)
    cell_number.value = next_candidate_number
    
    # Write the candidate name and link to the sheet
    cell_link = sheet.cell(row=starting_row, column=link_column)
    cell_link.value = user_name
    cell_link.hyperlink = user_link
    
    # Add the new candidate to the existing_candidates set
    existing_candidates.add(user_name)
    
    # Increment the candidate number for the next entry
    next_candidate_number += 1
    
    # Increment the starting_row for the next entry
    starting_row += 1
    last_used_row += 1
    
    # Save the workbook after each entry
    workbook.save("profiles.xlsx")


"""
#from openpyxl import Workbook
from openpyxl.styles import Font

from openpyxl import load_workbook

# Load the existing workbook
workbook = load_workbook("profiles.xlsx")

# Select the sheet from the workbook
sheet = workbook["Sheet"]
number_column = 1  # Column number for candidate numbers
link_column = 2   # Column number for candidate links

# Initialize an empty dictionary to store candidate names and links
existing_candidates = set()

# Populate the set with existing candidate names and links
for row in sheet.iter_rows(min_row=2, min_col=link_column, max_col=link_column, values_only=True):
    existing_candidates.add(row[0])
# Find the next available candidate number
number_values = [row[0].value for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=number_column, max_col=number_column) if row[0].value is not None]
next_candidate_number = max(number_values, default=0) + 1
# Find the last used row in the sheet
last_row = sheet.max_row

# Set the starting row for the next entry
current_row = last_row + 1
# Loop to collect candidate names and links
while True:
      # Ask the user for candidate name (exit loop if "exit" is entered)
    user_name = input("Please enter candidate name (enter 'exit' to quit): ")
    if user_name.lower() == "exit":
        break

    if user_name in existing_candidates:
      print("Candidate name already exists. Please enter a unique name.")
      continue
    
     # Ask the user for candidate link
    user_link = input("Please enter a candidate link: ")
    # Find the last used row with a hyperlink in the specified column
    if user_link in existing_candidates:
      print("Candidate link already exists. Please enter a unique link.")
      continue
    
    current_row = sheet.max_row + 1
    
    # Write the candidate number to the sheet
    cell_number = sheet.cell(row=current_row, column=number_column)
    cell_number.value = next_candidate_number
    
          # Write the candidate name and link to the sheet
    cell_link = sheet.cell(row=current_row, column=link_column)
    cell_link.value = user_name
    cell_link.hyperlink = user_link
    
      # Add the new candidate to the existing_candidates set
    existing_candidates.add(user_name)
    
        # Increment the candidate number for the next entry
    next_candidate_number += 1
      # Increment the current_row for the next entry
    current_row += 1
    last_row += 1
    
        # Save the workbook after each entry
    workbook.save("profiles.xlsx")
   
   """